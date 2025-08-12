from supplementary.readers.reader_interface import INPUT_TYPE, FormatReader
from supplementary.utils.data_fetching import get_extension, get_material_name, get_pmcid
from supplementary.utils.result_saver import get_output_type, save
from utils.logging.logging_setup import supplementary_error_logger
import io
import openpyxl
import pandas as pd

MAX_FILE_SIZE = 25000000  # 25 MB


def is_problematic_file(file_bytes, source):
    try:
        # Load the workbook from the BytesIO object
        workbook = openpyxl.load_workbook(file_bytes, read_only=True)

        # Iterate through each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get the first two rows (header and first data row)
            header_row = list(sheet.iter_rows(max_row=1, values_only=True))[0]
            first_data_row = list(
                sheet.iter_rows(min_row=2, max_row=2, values_only=True)
            )[0]
            # Compare the number of columns
            if len(first_data_row) > len(header_row) or len(first_data_row) > 15_000:
                return True  # Problematic file
        return False  # File is okay

    except Exception as e:
        error_message = "Excel file is problematic, to many columns to read."
        supplementary_error_logger.error("%s | %s | %s", source, error_message, str(e), exc_info=True)
        return True  # Treat as problematic if an error occurs


class EXCELFormatReader(FormatReader):
    def read_bytes(
        self,
        byte_contents,
        source: str,
        input_type=INPUT_TYPE.BYTES,
    ):
        if not byte_contents:
            return
        # if get_extension(source) == "xls":
        #     with TemporaryDirectory() as tempdir:
        #         xls_file = f"{tempdir}/temp.xlsx"
        #         xlsx_file = f"{tempdir}/temp.xlsx"
        #         with open(xls_file, "wb") as f:
        #             f.write(byte_contents)
        #         os.system(
        #             f"cd {tempdir} ; libreoffice --headless --convert-to xlsx temp.xls"
        #         )
        #         with open(xlsx_file, "rb") as f:
        #             xlsx_file = io.BytesIO(f.read())
        file_size = len(byte_contents)
        if file_size > MAX_FILE_SIZE:
            error_message = f"FILE TO LARGE -> {file_size=:_}\t {MAX_FILE_SIZE=:_}"
            supplementary_error_logger.error("%s | %s | %s", source, error_message, str(e), exc_info=True)
            return None
        xlsx_file = io.BytesIO(byte_contents)
        return self.process(xlsx_file, input_type, source)

    # def read_from_path(self, path, type=INPUT_TYPE.PATH):
    #     # TODO: Implement
    #     # return self.process(doc, type)
    #     pass

    def process(self, excel_file, type: INPUT_TYPE, source: str = ""):
        to_save = {}
        if is_problematic_file(excel_file, source):
            print(f"Skipping problematic file: {source}")
            return None
        try:
            df = pd.read_excel(excel_file, sheet_name=None)
        except Exception as e:
            error_message = "Error reading the excel file."
            supplementary_error_logger.error("%s | %s | %s", source, error_message, str(e), exc_info=True)
            return None
        for sheet_name, sheet in df.items():
            to_save[sheet_name] = sheet.to_string()

        if self.save:
            save(
                get_pmcid(source),
                get_material_name(source),
                to_save,
                get_output_type(get_extension(source)),
            )
        return to_save


if __name__ == "__main__":
    ...
